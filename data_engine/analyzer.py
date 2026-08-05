"""Core data-quality analysis for AuditData AI.

This module is intentionally independent from the web server and the UI.
That makes the "brain" reusable from a CLI, API, notebook, automation, or
future SaaS backend without rewriting the quality logic.
"""

from __future__ import annotations

import csv
import io
import logging
import re
import statistics
import zipfile
from collections import Counter
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


from .domain_rules import is_boolean_synonym, normalize_for_comparison
from .missing import is_missing

TYPE_THRESHOLD = 0.70

# AP-07: pesos configurables del overall. La exactitud estructural pesa más
# porque incluye errores de tipo; la completitud deja de eclipsar el resto.
QUALITY_WEIGHTS: dict[str, float] = {
    "completeness": 0.30,
    "consistency": 0.20,
    "accuracy": 0.35,
    "uniqueness": 0.15,
}

DELIMITER_NAMES: dict[str, str] = {"comma": ",", "semicolon": ";", "tab": "\t", "pipe": "|"}
ENCODINGS_TO_TRY: list[str] = ["utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"]



@dataclass
class ColumnProfile:
    """Analysis result for one dataset column."""

    name: str
    detected_type: str
    total_rows: int
    missing: int
    unique_values: int
    examples: list[str] = field(default_factory=list)
    format_issues: int = 0
    format_groups: list[dict[str, Any]] = field(default_factory=list)
    outliers: int = 0
    outlier_examples: list[Any] = field(default_factory=list)
    min_value: float | None = None
    max_value: float | None = None
    mean: float | None = None
    median: float | None = None
    distribution_pct: float = 0.0
    value_distribution: list[dict[str, Any]] = field(default_factory=list)
    invalid_type_count: int = 0
    outlier_analysis_skipped: bool = False


def load_dataset(filename: str, payload: bytes, delimiter: str | None = None, encoding: str | None = None, header_row: int | None = None) -> tuple[list[str], list[dict[str, Any]], int]:
    """Load CSV or XLSX bytes into headers and row dictionaries.

    Returns (headers, rows, header_row_index) where header_row_index is the
    0-based position of the header row in the original file (used to compute
    the correct Excel row offset for diagnostics).

    delimiter/encoding/header_row only apply to CSV files; they come from the
    file preview settings so the analysis uses the same parsing as the preview.
    """

    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return _load_csv(payload, delimiter=delimiter, encoding=encoding, header_row=header_row)
    if lowered.endswith((".xlsx", ".xlsm")):
        return _load_xlsx(payload)
    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


def analyze_dataset(filename: str, payload: bytes, duplicate_key_columns: list[str] | None = None, delimiter: str | None = None, encoding: str | None = None, header_row: int | None = None) -> dict[str, Any]:
    """Run the complete reusable quality diagnosis for a dataset."""

    headers, rows, _header_idx = load_dataset(filename, payload, delimiter=delimiter, encoding=encoding, header_row=header_row)
    return _analyze_rows(headers, rows, filename, duplicate_key_columns=duplicate_key_columns)


def _analyze_rows(headers: list[str], rows: list[dict[str, Any]], filename: str, duplicate_key_columns: list[str] | None = None, frozen_outlier_bounds: dict[str, tuple[float, float]] | None = None) -> dict[str, Any]:
    """Core analysis from in-memory headers/rows (CL-08): re-perfila sin re-parsear CSV.

    frozen_outlier_bounds (CL-10): límites IQR del dataset ORIGINAL por columna,
    para que el re-perfilado post-limpieza no recalcule outliers sobre datos ya
    imputados/corregidos (evita que una acción correctiva degrade la exactitud).
    """
    duplicate_rows = _count_duplicate_rows(headers, rows, key_columns=duplicate_key_columns)
    columns = [_profile_column(header, rows, outlier_bounds=(frozen_outlier_bounds or {}).get(header)) for header in headers]
    scores = _quality_scores(columns, duplicate_rows, len(rows), max(len(headers), 1))

    return {
        "filename": filename,
        "generated_at": datetime.now(timezone(timedelta(hours=-5))).isoformat(timespec="seconds"),
        "row_count": len(rows),
        "column_count": len(headers),
        "headers": headers,
        "duplicate_rows": duplicate_rows,
        "scores": scores,
        "columns": [profile.__dict__ for profile in columns],
        "recommendations": _recommendations(columns, duplicate_rows),
        "preview": rows[:10],
    }


def apply_cleaning_actions(filename: str, payload: bytes, actions: list[dict[str, Any]], duplicate_key_columns: list[str] | None = None, delimiter: str | None = None, encoding: str | None = None, header_row: int | None = None) -> dict[str, Any]:
    """Apply documented cleaning actions and return before/after evidence.

    Every action creates a log entry. This follows the professional rule from
    the source lessons: cleaning is not complete until the decisión is traceable.
    Now also tracks cell-level changes for the audit log.
    """

    headers, rows, header_idx = load_dataset(filename, payload, delimiter=delimiter, encoding=encoding, header_row=header_row)
    before = analyze_dataset(filename, payload, duplicate_key_columns=duplicate_key_columns, delimiter=delimiter, encoding=encoding, header_row=header_row)
    # CL-10: congelar los límites IQR del dataset ORIGINAL para el re-perfilado
    # post-limpieza. Imputar/rellenar no debe recalcular el rango de outliers
    # sobre datos ya corregidos (una acción correctiva nunca degrada exactitud).
    frozen_outlier_bounds: dict[str, tuple[float, float]] = {}
    for header in headers:
        parsed = _to_float_column([_normalize_missing(row.get(header, "")) for row in rows])
        numeric = [value for value in parsed if value is not None]
        bounds = _numeric_outlier_bounds(numeric)
        if bounds is not None:
            frozen_outlier_bounds[header] = bounds
    log: list[dict[str, str]] = []
    changelog: list[dict[str, Any]] = []

    justification_params = [
        (a.get("column", "") or "Dataset", a.get("kind"), a.get("reason", "").strip() or "Decision registrada sin detalle adicional.")
        for a in actions
    ]
    from data_engine.ai_advisor import get_justifications_batch
    justifications = get_justifications_batch(justification_params)

    for i, action in enumerate(actions):
        kind = action.get("kind")
        column = action.get("column", "")
        target_rows = action.get("rows")  # Excel row numbers from frontend
        if target_rows is not None:
            # CL-05: convertir Excel row -> índice 0-based de datos usando el header REAL.
            target_rows = [r - header_idx - 2 for r in target_rows]

        ai_reason = justifications[i]

        if kind == "delete_column" and column in headers:
            changelog.append({
                "action": "Eliminar columna",
                "column": column,
                "reason": ai_reason,
                "changes": [{"row": "TODAS", "column": column, "old": "(valor presente)", "new": "(columna eliminada)"}],
            })
            headers.remove(column)
            for row in rows:
                row.pop(column, None)
            log.append(_log_entry(column, "Eliminar columna", ai_reason, "Columna eliminada del dataset limpio."))

        elif kind == "drop_missing_rows" and column in headers:
            before_count = len(rows)
            dropped_rows = []
            kept_rows = []
            target_set = set(target_rows) if target_rows is not None else None
            for idx, row in enumerate(rows):
                is_missing = _normalize_missing(row.get(column, "")) == ""
                in_target = target_set is None or idx in target_set
                if is_missing and in_target:
                    dropped_rows.append(dict(row))
                else:
                    kept_rows.append(row)
            rows = kept_rows
            for dr in dropped_rows:
                row_id = dr.get("id", dr.get("ID", dr.get(headers[0], "?")))
                changelog.append({
                    "action": "Eliminar fila",
                    "column": column,
                    "reason": ai_reason,
                    "changes": [{"row": str(row_id), "column": column, "old": "(vacío)", "new": "(fila eliminada)"}],
                })
            scope = "filas especificas" if target_rows else "todas las filas"
            log.append(_log_entry(column, "Eliminar filas con faltantes", ai_reason, f"{before_count - len(rows)} filas eliminadas ({scope})."))

        elif kind == "impute_missing" and column in headers:
            method = action.get("method", "mode")
            value = _imputation_value(rows, column, method, action.get("value"))
            target_set = set(target_rows) if target_rows is not None else None
            changed, entries = _apply_fill_cells(
                rows, headers, column, target_set,
                lambda v: _normalize_missing(v) == "",
                value, f"Imputar con {method}", ai_reason,
            )
            changelog.extend(entries)
            scope = "filas especificas" if target_rows else "todas las vacias"
            log.append(_log_entry(column, f"Imputar faltantes con {method}", ai_reason, f"{changed} valores reemplazados ({scope})."))

        elif kind == "fill_missing" and column in headers:
            method = action.get("method", "mode")
            target_set = set(target_rows) if target_rows is not None else None
            if method == "null":
                changed, entries = _apply_fill_cells(
                    rows, headers, column, target_set,
                    lambda v: _normalize_missing(v) == "",
                    "NULL", "Rellenar con NULL", ai_reason,
                )
                changelog.extend(entries)
                log.append(_log_entry(column, "Rellenar vacíos con NULL", ai_reason, f"{changed} celdas rellenadas con NULL."))
            else:
                value = _imputation_value(rows, column, method, action.get("value"))
                changed, entries = _apply_fill_cells(
                    rows, headers, column, target_set,
                    lambda v: _normalize_missing(v) == "",
                    value, f"Rellenar con {method}", ai_reason,
                )
                changelog.extend(entries)
                log.append(_log_entry(column, f"Rellenar vacíos con {method}", ai_reason, f"{changed} celdas rellenadas ({method})."))

        elif kind == "standardize_text" and column in headers:
            mode = action.get("method", "title")
            changed = 0
            target_set = set(target_rows) if target_rows is not None else None
            for idx, row in enumerate(rows):
                in_target = target_set is None or idx in target_set
                if not in_target:
                    continue
                original = str(row.get(column, ""))
                updated = _standardize_text(original, mode)
                if original != updated:
                    row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                    changelog.append({
                        "action": f"Estandarizar ({mode})",
                        "column": column,
                        "reason": ai_reason,
                        "changes": [{"row": str(row_id), "column": column, "old": original, "new": updated}],
                    })
                    row[column] = updated
                    changed += 1
            scope = "filas especificas" if target_rows else "todas las filas"
            log.append(_log_entry(column, f"Estandarizar texto ({mode})", ai_reason, f"{changed} celdas normalizadas ({scope})."))

        elif kind == "remove_duplicate_rows":
            before_count = len(rows)
            target_set = set(target_rows) if target_rows is not None else None
            if target_set is not None:
                clean_rows = []
                for idx, row in enumerate(rows):
                    if idx in target_set:
                        row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                        changelog.append({
                            "action": "Eliminar duplicado (seleccionado)",
                            "column": "Dataset",
                            "reason": ai_reason,
                            "changes": [{"row": str(row_id), "column": "*", "old": "(fila duplicada)", "new": "(fila eliminada)"}],
                        })
                    else:
                        clean_rows.append(row)
                rows = clean_rows
                log.append(_log_entry("Dataset", "Eliminar filas duplicadas", ai_reason, f"{before_count - len(rows)} filas eliminadas (seleccionadas)."))
            else:
                seen: set[tuple[str, ...]] = set()
                clean_rows = []
                compare_headers = duplicate_key_columns if duplicate_key_columns else headers
                for row in rows:
                    key = tuple(normalize_for_comparison(row.get(h, "")) for h in compare_headers)
                    if key not in seen:
                        seen.add(key)
                        clean_rows.append(row)
                    else:
                        row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                        changelog.append({
                            "action": "Eliminar duplicado",
                            "column": "Dataset",
                            "reason": ai_reason,
                            "changes": [{"row": str(row_id), "column": "*", "old": "(fila duplicada)", "new": "(fila eliminada)"}],
                        })
                rows = clean_rows
                log.append(_log_entry("Dataset", "Eliminar filas duplicadas", ai_reason, f"{before_count - len(rows)} filas duplicadas eliminadas."))

        elif kind == "flag_outliers" and column in headers:
            if target_rows:
                target_set = set(target_rows)
                marked = 0
                for idx, row in enumerate(rows):
                    if idx not in target_set:
                        continue
                    row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                    changelog.append({
                        "action": "Marcar outlier",
                        "column": column,
                        "reason": ai_reason,
                        "changes": [{"row": str(row_id), "column": column, "old": str(row.get(column, "")), "new": "(marcado para revision)"}],
                    })
                    marked += 1
                log.append(_log_entry(column, "Marcar outliers para revision", ai_reason, f"{marked} filas marcadas para revision manual."))
            else:
                outlier_indices = _outlier_row_indices(rows, column)
                if outlier_indices:
                    for idx in outlier_indices:
                        row = rows[idx]
                        row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                        changelog.append({
                            "action": "Marcar outlier",
                            "column": column,
                            "reason": ai_reason,
                            "changes": [{"row": str(row_id), "column": column, "old": str(row.get(column, "")), "new": "(marcado para revision)"}],
                        })
                    log.append(_log_entry(column, "Marcar outliers para revision", ai_reason, f"{len(outlier_indices)} filas atípicas detectadas y marcadas para revisión manual."))
                else:
                    log.append(_log_entry(column, "Marcar outliers para revision", ai_reason, "Sin cambios destructivos en el dataset."))

        elif kind == "replace_with_null" and column in headers:
            changed = 0
            target_set = set(target_rows) if target_rows is not None else None
            for idx, row in enumerate(rows):
                in_target = target_set is None or idx in target_set
                if not in_target:
                    continue
                old_val = row.get(column, "")
                if old_val:
                    row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                    changelog.append({
                        "action": "Reemplazar con NULL",
                        "column": column,
                        "reason": ai_reason,
                        "changes": [{"row": str(row_id), "column": column, "old": str(old_val), "new": "(null)"}],
                    })
                    row[column] = ""
                    changed += 1
            scope = "filas especificas" if target_rows else "todas las filas"
            log.append(_log_entry(column, "Reemplazar valores con NULL", ai_reason, f"{changed} valores reemplazados ({scope})."))

        elif kind == "rename_column" and column in headers:
            new_name = action.get("value", "").strip()
            if new_name and new_name not in headers:
                idx = headers.index(column)
                headers[idx] = new_name
                for row in rows:
                    if column in row:
                        row[new_name] = row.pop(column)
                changelog.append({
                    "action": f"Renombrar a '{new_name}'",
                    "column": column,
                    "reason": ai_reason,
                    "changes": [{"row": "TODAS", "column": column, "old": column, "new": new_name}],
                })
                log.append(_log_entry(column, f"Renombrar a '{new_name}'", ai_reason, f"Columna renombrada a '{new_name}'."))

        elif kind == "replace_value" and column in headers:
            target_val = action.get("method", "")
            repl_val = action.get("value", "")
            changed = 0
            for row in rows:
                if str(row.get(column, "")) == target_val:
                    row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                    changelog.append({
                        "action": f"Reemplazar '{target_val}' -> '{repl_val}'",
                        "column": column,
                        "reason": ai_reason,
                        "changes": [{"row": str(row_id), "column": column, "old": target_val, "new": repl_val}],
                    })
                    row[column] = repl_val
                    changed += 1
            log.append(_log_entry(column, f"Reemplazar '{target_val}' con '{repl_val}'", ai_reason, f"{changed} valores reemplazados."))

        elif kind == "change_type" and column in headers:
            target_type = action.get("value", "text")
            changed = 0
            comma_is = _classify_numeric_separators([str(row.get(column, "")) for row in rows])
            for row in rows:
                val = row.get(column, "")
                if target_type == "number":
                    casted = _to_float(val, separator_mode=comma_is)
                    if casted is not None:
                        row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                        changelog.append({
                            "action": f"Tipo: {target_type}",
                            "column": column,
                            "reason": ai_reason,
                            "changes": [{"row": str(row_id), "column": column, "old": str(val), "new": str(casted)}],
                        })
                        row[column] = casted
                        changed += 1
                elif target_type == "text":
                    row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                    changelog.append({
                        "action": f"Tipo: {target_type}",
                        "column": column,
                        "reason": ai_reason,
                        "changes": [{"row": str(row_id), "column": column, "old": str(val), "new": str(val)}],
                    })
                    row[column] = str(val)
                    changed += 1
                elif target_type == "boolean":
                    syn = is_boolean_synonym(str(val))
                    new_val = "si" if syn is True else "no"
                    row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
                    changelog.append({
                        "action": f"Tipo: {target_type}",
                        "column": column,
                        "reason": ai_reason,
                        "changes": [{"row": str(row_id), "column": column, "old": str(val), "new": new_val}],
                    })
                    row[column] = new_val
                    changed += 1
            log.append(_log_entry(column, f"Cambiar tipo de dato a {target_type}", ai_reason, f"{changed} valores convertidos."))

        elif kind == "fill_empty" and column in headers:
            fill_val = action.get("value", "NULL")
            target_set = set(target_rows) if target_rows is not None else None
            changed, entries = _apply_fill_cells(
                rows, headers, column, target_set,
                lambda v: not v or not str(v).strip(),
                fill_val, f"Rellenar vacío con '{fill_val}'", ai_reason,
            )
            changelog.extend(entries)
            scope = "filas especificas" if target_rows else "todas las vacias"
            log.append(_log_entry(column, f"Rellenar vacíos con '{fill_val}'", ai_reason, f"{changed} celdas rellenadas ({scope})."))

        elif kind == "review_issue":
            changelog.append({
                "action": "Revision manual",
                "column": column,
                "reason": ai_reason,
                "changes": [],
            })
            log.append(_log_entry(column, "Revision manual", ai_reason, "Issue revisado sin modificar datos."))

    clean_csv = rows_to_csv(headers, rows)
    # CL-08: after re-perfilado en memoria, sin re-parsear el CSV limpio.
    # CL-10: con límites de outliers congelados del dataset original.
    after = _analyze_rows(headers, rows, _clean_filename(filename), duplicate_key_columns=duplicate_key_columns, frozen_outlier_bounds=frozen_outlier_bounds)
    return {"before": before, "after": after, "actions": log, "clean_csv": clean_csv, "changelog": changelog}



def build_markdown_report(analysis: dict[str, Any], analyst: str = "-", version: str = "v1.0") -> str:
    """Create a professional Markdown report from an analysis object."""

    scores = analysis["scores"]
    lines = [
        f"# Data Cleaning Report - {analysis['filename']}",
        "",
        "## Información general",
        f"- Dataset: {analysis['filename']}",
        f"- Analista: {analyst or '-'}",
        f"- Version del informe: {version or 'v1.0'}",
        f"- Fecha técnica: {analysis['generated_at']}",
        f"- Filas: {analysis['row_count']}",
        f"- Columnas: {analysis['column_count']}",
        "",
        "## Resumen ejecutivo",
        _executive_summary(analysis),
        "",
        "## Indicadores clave",
        "| Indicador | Resultado |",
        "|---|---:|",
        f"| Completitud | {scores['completeness']}% |",
        f"| Consistencia | {scores['consistency']}% |",
        f"| Exactitud estructural | {scores['accuracy']}% |",
        f"| Unicidad | {scores['uniqueness']}% |",
        f"| Calidad general | {scores['overall']}% |",
        f"| Filas duplicadas | {analysis['duplicate_rows']} |",
        "",
        "## Problemas encontrados",
        "| Columna | Tipo | Faltantes | Inconsistencias | Outliers |",
        "|---|---|---:|---:|---:|",
    ]

    for column in analysis["columns"]:
        lines.append(
            f"| {column['name']} | {column['detected_type']} | {column['missing']} | "
            f"{column['format_issues']} | {column['outliers']} |"
        )

    lines.extend(["", "## Recomendaciones"])
    for item in analysis["recommendations"]:
        lines.append(f"- **{item['priority']}** - {item['message']}")

    lines.extend(
        [
            "",
            "## Criterio metodologico",
            "La herramienta no inventa datos. Los hallazgos se calculan sobre el dataset recibido y las acciones de limpieza deben validarse con criterio de negocio antes de reemplazar, eliminar o imputar valores.",
        ]
    )
    return "\n".join(lines) + "\n"


def _format_action_markdown(item: dict[str, Any]) -> tuple[str, str]:
    """Convert an action dict to (label, result) strings for markdown report."""
    kind = item.get("kind", item.get("action", ""))
    column = item.get("column", "-")
    method = item.get("method", "")
    value = item.get("value", "")
    rows = item.get("rows", [])

    _labels = {
        "analyst_note": "Nota del Analista",
        "delete_column": "Eliminar columna",
        "drop_missing_rows": "Eliminar filas con faltantes",
        "impute_missing": "Imputar faltantes",
        "standardize_text": "Estandarizar texto",
        "remove_duplicate_rows": "Eliminar duplicados",
        "flag_outliers": "Marcar outliers",
        "fill_missing": "Rellenar celdas vacias",
        "fill_empty": "Rellenar celdas vacias",
        "replace_with_null": "Reemplazar con NULL",
        "rename_column": "Renombrar columna",
        "drop_duplicates": "Eliminar duplicados",
        "convert_type": "Convertir tipo de dato",
        "change_type": "Cambiar tipo de dato",
        "drop_rows": "Eliminar filas",
        "fix_format": "Corregir formato",
        "replace_value": "Reemplazar valor",
        "review_issue": "Revision manual",
    }
    label = _labels.get(kind, kind) or "-"

    if kind in ("fill_empty", "fill_missing"):
        result = f"Se rellenaron celdas vacias con '{value}'" if value else "Celdas vacias rellenadas"
    elif kind == "replace_value":
        result = f"Reemplazado: {method} -> {value}" if method else f"Valores reemplazados con '{value}'"
    elif kind == "rename_column":
        result = f"Renombrada a '{value}'" if value else "Columna renombrada"
    elif kind in ("change_type", "convert_type"):
        result = f"Tipo cambiado a '{method}'" if method else "Tipo de dato cambiado"
    elif kind == "delete_column":
        result = f"Columna '{column}' eliminada del dataset"
    elif kind in ("remove_duplicate_rows", "drop_duplicates"):
        count = len(rows) if rows else 0
        result = f"{count} filas duplicadas eliminadas" if count else "Filas duplicadas eliminadas"
    elif kind == "drop_missing_rows":
        count = len(rows) if rows else 0
        result = f"{count} filas con valores faltantes eliminadas" if count else "Filas con faltantes eliminadas"
    elif kind == "drop_rows":
        count = len(rows) if rows else 0
        result = f"{count} filas eliminadas" if count else "Filas eliminadas"
    elif kind == "analyst_note":
        result = "Registrado en bitacora"
    elif kind == "standardize_text":
        result = f"Texto normalizado ({method})" if method else "Texto normalizado"
    elif kind == "fix_format":
        result = f"Formato corregido ({method})" if method else "Formato corregido"
    elif kind == "flag_outliers":
        result = f"{len(rows)} valores atipicos marcados" if rows else "Valores atipicos marcados"
    else:
        result = method or "Accion aplicada"

    return label, result


def build_cleaning_markdown_report(cleaning: dict[str, Any], analyst: str = "-", version: str = "v1.0", row_meaning: str = "", analysis_objective: str = "") -> str:
    """Create a comprehensive before/after report matching the academic Data Cleaning Report standard."""

    before = cleaning["before"]
    after = cleaning["after"]
    actions = cleaning.get("actions", [])

    context_lines = []
    if row_meaning:
        context_lines.append(f"- Que representa cada fila: {row_meaning}")
    if analysis_objective:
        context_lines.append(f"- Objetivo del análisis: {analysis_objective}")

    lines = [
        f"# Data Cleaning Report - {before['filename']}",
        "",
        "## 1. Información General",
        f"- Dataset original: {before['filename']}",
        f"- Dataset limpio: {after['filename']}",
        f"- Analista: {analyst or '-'}",
        f"- Version del informe: {version or 'v1.0'}",
    ] + context_lines + [
        f"- Registros antes: {before['row_count']}",
        f"- Registros después: {after['row_count']}",
        f"- Columnas antes: {before['column_count']}",
        f"- Columnas después: {after['column_count']}",
        f"- Acciones documentadas: {len(actions)}",
        f"- Fecha de generacion: {after['generated_at']}",
        "- Herramienta utilizada: AuditData AI - Motor Python",
        "",
        "## 2. Resumen Ejecutivo",
        _cleaning_resumen(before, after, actions),
        "",
        "## 3. Indicadores Clave del Dataset",
        "| Indicador | Antes | Después |",
        "|---|---:|---:|",
        f"| Registros | {before['row_count']} | {after['row_count']} |",
        f"| Columnas | {before['column_count']} | {after['column_count']} |",
        f"| Filas duplicadas | {before['duplicate_rows']} | {after['duplicate_rows']} |",
        f"| Completitud | {before['scores']['completeness']}% | {after['scores']['completeness']}% |",
        f"| Consistencia | {before['scores']['consistency']}% | {after['scores']['consistency']}% |",
        f"| Exactitud estructural | {before['scores']['accuracy']}% | {after['scores']['accuracy']}% |",
        f"| Unicidad | {before['scores']['uniqueness']}% | {after['scores']['uniqueness']}% |",
        f"| Calidad general | {before['scores']['overall']}% | {after['scores']['overall']}% |",
        "",
        "## 4. Problemas Encontrados",
    ]

    missing_before = [c for c in before["columns"] if c.get("missing", 0) > 0]
    format_before = [c for c in before["columns"] if c.get("format_issues", 0) > 0]
    outliers_before = [c for c in before["columns"] if c.get("outliers", 0) > 0]

    lines.append("")
    lines.append("### 4.1 Valores Faltantes por Columna")
    if missing_before:
        total_missing = sum(c["missing"] for c in missing_before)
        lines.append(f"El dataset presento {total_missing} celdas vacias:")
        lines.append("| Columna | Faltantes | % Columna | Tipo detectado |")
        lines.append("|---|---:|---:|---|")
        for c in missing_before:
            pct = round(c["missing"] / max(before["row_count"], 1) * 100, 1)
            lines.append(f"| {c['name']} | {c['missing']} | {pct}% | {c['detected_type']} |")
    else:
        lines.append("No se detectaron valores faltantes.")

    lines.append("")
    lines.append("### 4.2 Filas Duplicadas")
    if before["duplicate_rows"] > 0:
        dup_pct = round(before["duplicate_rows"] / max(before["row_count"], 1) * 100, 1)
        lines.append(f"Se detectaron {before['duplicate_rows']} filas duplicadas ({dup_pct}% del total).")
    else:
        lines.append("No se detectaron filas duplicadas.")

    lines.append("")
    lines.append("### 4.3 Errores de Escritura y Variantes de Texto")
    if format_before:
        total_format = sum(c["format_issues"] for c in format_before)
        lines.append(f"Se encontraron {total_format} inconsistencias de formato en {len(format_before)} columnas.")
        for col in format_before:
            groups = col.get("format_groups", [])
            if groups:
                lines.append(f"**{col['name']}:**")
                for g in groups[:5]:
                    variants = g.get("variants", [])
                    lines.append(f"  - Canonical: '{g.get('canonical', '')}' | Variantes: {', '.join(variants)}")
    else:
        lines.append("No se detectaron errores de escritura significativos.")

    lines.append("")
    lines.append("### 4.4 Formatos Inconsistentes")
    cols_with_groups = [c for c in before["columns"] if c.get("format_groups")]
    if cols_with_groups:
        lines.append(f"{len(cols_with_groups)} columnas presentan formatos mixtos.")
    else:
        lines.append("No se detectaron formatos inconsistentes.")

    lines.append("")
    lines.append("### 4.5 Categorías Inconsistentes")
    if format_before:
        lines.append(f"{len(format_before)} columnas presentan categorías fragmentadas.")
    else:
        lines.append("No se detectaron categorías inconsistentes.")

    lines.append("")
    lines.append("### 4.6 Valores Atipicos")
    if outliers_before:
        total_outliers = sum(c["outliers"] for c in outliers_before)
        lines.append(f"Se detectaron {total_outliers} valores atipicos en {len(outliers_before)} columnas.")
        actions_by_col = {}
        for a in actions:
            col = a.get("column", "")
            if col not in actions_by_col:
                actions_by_col[col] = a.get("action", "")
        lines.append("| Columna | Outliers | Ejemplos | Acción |")
        lines.append("|---|---:|---|---|")
        for c in outliers_before:
            examples = ", ".join(str(v) for v in (c.get("outlier_examples") or [])[:3])
            lines.append(f"| {c['name']} | {c['outliers']} | {examples} | {actions_by_col.get(c['name'], 'Sin acción')} |")
    else:
        lines.append("No se detectaron valores atipicos.")

    lines.extend(["", "## 5. Plan y Acciones de Limpieza"])
    if actions:
        lines.append(f"Se documentaron {len(actions)} acciones de limpieza.")
        lines.append("| N. | Columna | Acción | Justificación | Resultado |")
        lines.append("|---|---|---|---|---|")
        for i, item in enumerate(actions):
            action_label, result_text = _format_action_markdown(item)
            lines.append(f"| {i+1} | {item.get('column', '-')} | {action_label} | {item.get('reason', '-')} | {result_text} |")
    else:
        lines.append("No se aplicaron acciones de limpieza.")

    lines.extend(["", "## 6. Evaluación de Calidad - Antes vs Después"])
    dims = [("Completitud", "completeness"), ("Consistencia", "consistency"), ("Exactitud estructural", "accuracy"), ("Unicidad", "uniqueness")]
    lines.append("| Dimension | Antes | Después | Cambio |")
    lines.append("|---|---|---|---|")
    for label, key in dims:
        diff = round(after["scores"][key] - before["scores"][key], 2)
        sign = "+" if diff > 0 else ""
        lines.append(f"| {label} | {before['scores'][key]}% | {after['scores'][key]}% | {sign}{diff}% |")
    overall_diff = round(after["scores"]["overall"] - before["scores"]["overall"], 2)
    overall_sign = "+" if overall_diff > 0 else ""
    lines.append(f"| Calidad general | {before['scores']['overall']}% | {after['scores']['overall']}% | {overall_sign}{overall_diff}% |")

    lines.extend(["", "## 7. Checklist de Validación Final"])
    checks = [
        ("Completitud >= 95%", after["scores"]["completeness"] >= 95),
        ("Consistencia >= 95%", after["scores"]["consistency"] >= 95),
        ("Exactitud estructural >= 95%", after["scores"]["accuracy"] >= 95),
        ("Sin duplicados pendientes", after["duplicate_rows"] == 0),
        ("Acciones documentadas", len(actions) > 0),
        ("Calidad general >= 90%", after["scores"]["overall"] >= 90),
    ]
    lines.append("| Criterio | Estado |")
    lines.append("|---|---|")
    for criterion, passed in checks:
        lines.append(f"| {criterion} | {'Cumple' if passed else 'Requiere revision'} |")

    lines.extend(["", "## 8. Riesgos Identificados"])
    risks = _risk_list(after)
    for risk in risks:
        lines.append(f"- {risk}")

    lines.extend(["", "## 9. Metodología de Calculo"])
    lines.append("**Completitud:** 100% - (celdas vacias / total de celdas) * 100.")
    lines.append("**Consistencia:** 100% - (inconsistencias de formato / total de celdas) * 100.")
    lines.append("**Exactitud estructural:** 100% - ((errores de tipo + valores atipicos) / total de celdas) * 100. Atipicos calculados con IQR.")
    lines.append("**Unicidad:** 100% - (filas duplicadas / total de filas) * 100.")
    lines.append("**Calidad general:** Promedio ponderado (completitud 30%, consistencia 20%, exactitud estructural 35%, unicidad 15%).")

    lines.extend(["", "## 10. Conclusión Final"])
    lines.append(_conclusión(after))

    return "\n".join(lines) + "\n"


def rows_to_csv(headers: list[str], rows: list[dict[str, Any]]) -> str:
    """Serialize current dataset rows into CSV for download or re-analysis."""

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header, "") for header in headers})
    return output.getvalue()


def csv_to_xlsx(csv_content: str, filename: str = "dataset.xlsx") -> bytes:
    """Convert CSV string to XLSX bytes using openpyxl."""
    if not openpyxl:
        raise RuntimeError("openpyxl no esta instalado. Instala con: pip install openpyxl")
    reader = csv.DictReader(io.StringIO(csv_content))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Limpios"
    headers = reader.fieldnames or []
    ws.append(headers)
    for row in reader:
        ws.append([row.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_header_row(raw_rows: list[tuple]) -> int:
    """Detect which row is the actual header row, skipping title/metadata rows.

    A row is considered a header row when it has at least 3 non-empty cells
    and at least 2 distinct non-empty values.  Title rows typically have a
    single merged cell and the rest are None.
    """
    for idx, row in enumerate(raw_rows[:5]):
        non_empty = [str(v).strip() for v in row if v is not None and str(v).strip()]
        distinct = set(non_empty)
        if len(non_empty) >= 3 and len(distinct) >= 2:
            return idx
    return 0


def _decode_text_with_encoding(payload: bytes, encoding: str | None = None) -> tuple[str, str]:
    """Decode CSV bytes trying the requested encoding first, then known ones.

    Returns (text, detected_encoding_name).
    """
    candidates = [encoding] if encoding else []
    candidates.extend(ENCODINGS_TO_TRY)
    for enc in candidates:
        if not enc:
            continue
        try:
            return payload.decode(enc), enc.replace("-sig", "")
        except (UnicodeDecodeError, LookupError):
            continue
    # latin-1 nunca falla: último recurso sin romper el análisis
    return payload.decode("latin-1", errors="replace"), "latin-1"


def _decode_text(payload: bytes, encoding: str | None = None) -> str:
    text, _ = _decode_text_with_encoding(payload, encoding=encoding)
    return text


def _resolve_delimiter(prefer: str | None) -> str | None:
    """Map a delimiter name (comma/semicolon/tab/pipe) or char to its character."""
    if not prefer:
        return None
    if prefer in DELIMITER_NAMES:
        return DELIMITER_NAMES[prefer]
    if prefer in DELIMITER_NAMES.values():
        return prefer
    return None


def _count_outside_quotes(lines: list[str], ch: str) -> int:
    """Count occurrences of `ch` outside of double-quoted sections.

    Quote state is carried across line boundaries so a quoted field that spans
    multiple lines is handled correctly.
    """
    count = 0
    in_quotes = False
    for line in lines:
        i = 0
        n = len(line)
        while i < n:
            c = line[i]
            if c == '"':
                if in_quotes and i + 1 < n and line[i + 1] == '"':
                    i += 2
                    continue
                in_quotes = not in_quotes
            elif c == ch and not in_quotes:
                count += 1
            i += 1
    return count


def _split_csv_line(line: str, delim: str) -> list[str]:
    """Split a CSV line respecting quoted fields."""
    try:
        reader = csv.reader([line], delimiter=delim)
        return next(reader)
    except (csv.Error, StopIteration):
        return line.split(delim)


def _detect_delimiter(lines: list[str], prefer: str | None = None) -> str:
    """Auto-detect the CSV delimiter counting candidates OUTSIDE quotes.

    A `prefer` (name or char) wins over heuristics when provided.
    """
    resolved = _resolve_delimiter(prefer)
    if resolved is not None:
        return resolved

    sample = lines[:20]
    counts: dict[str, int] = {}
    for name, ch in DELIMITER_NAMES.items():
        total = _count_outside_quotes(sample, ch)
        if total > 0:
            counts[name] = total
    if not counts:
        return ","
    return DELIMITER_NAMES[max(counts, key=counts.get)]


def _find_header_index(lines: list[str], delim: str) -> int:
    """Find the 0-based header row among the first 5 lines.

    A row is considered a header row when it has at least 3 non-empty fields.
    """
    for i, line in enumerate(lines[:5]):
        fields = [f.strip() for f in _split_csv_line(line, delim) if f.strip()]
        if len(fields) >= 3:
            return i
    return 0


def _load_csv(payload: bytes, delimiter: str | None = None, encoding: str | None = None, header_row: int | None = None) -> tuple[list[str], list[dict[str, Any]], int]:
    """Load CSV bytes into headers and row dictionaries.

    delimiter: name ("comma"/"semicolon"/"tab"/"pipe") or delimiter character.
    encoding: name to try first.
    header_row: 0-based row index of the header; None auto-detects.
    """
    text = _decode_text(payload, encoding=encoding)
    lines = text.splitlines()
    if not lines:
        return [], [], 0

    delim = _detect_delimiter(lines, prefer=delimiter)
    header_idx = header_row if header_row is not None else _find_header_index(lines, delim)

    trimmed = "\n".join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(trimmed), delimiter=delim)
    headers = reader.fieldnames or []
    rows = [{header: row.get(header, "") for header in headers} for row in reader]
    return headers, rows, header_idx


def _load_xlsx(payload: bytes) -> tuple[list[str], list[dict[str, Any]], int]:
    if openpyxl is None:
        raise ValueError("openpyxl no esta disponible para leer XLSX.")
    if not zipfile.is_zipfile(io.BytesIO(payload)):
        raise ValueError("El archivo XLSX no parece valido.")

    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        return [], [], 0

    header_idx = _find_header_row(raw_rows)

    headers = [str(value).strip() if value is not None else f"columna_{i+1}" for i, value in enumerate(raw_rows[header_idx])]
    rows = []
    for raw in raw_rows[header_idx + 1:]:
        row = {}
        for index, header in enumerate(headers):
            row[header] = raw[index] if index < len(raw) and raw[index] is not None else ""
        rows.append(row)
    return headers, rows, header_idx


def detect_file_settings(filename: str, payload: bytes) -> dict[str, Any]:
    """Detect encoding, delimiter, header row and return a preview of the file."""
    lowered = filename.lower()
    is_xlsx = lowered.endswith((".xlsx", ".xlsm"))

    if is_xlsx:
        if openpyxl is None:
            return {"error": "openpyxl no disponible"}
        if not zipfile.is_zipfile(io.BytesIO(payload)):
            return {"error": "Archivo XLSX no valido"}
        workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            return {"error": "Archivo vacío"}
        header_idx = _find_header_row(raw_rows)
        headers = [str(v).strip() if v is not None else f"col_{i+1}" for i, v in enumerate(raw_rows[header_idx])]
        preview = []
        for raw in raw_rows[header_idx + 1:header_idx + 11]:
            row = {}
            for i, h in enumerate(headers):
                row[h] = raw[i] if i < len(raw) and raw[i] is not None else ""
            preview.append(row)
        return {
            "encoding": "utf-8",
            "delimiter": ",",
            "detected_header_row": header_idx,
            "total_rows": len(raw_rows) - header_idx - 1,
            "headers": headers,
            "preview": preview,
            "format": "xlsx",
        }

    text, detected_encoding = _decode_text_with_encoding(payload)

    lines = text.splitlines()
    if not lines:
        return {"error": "Archivo vacío"}

    delim_char = _detect_delimiter(lines)
    header_idx = _find_header_index(lines, delim_char)

    trimmed = "\n".join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(trimmed), delimiter=delim_char)
    headers = reader.fieldnames or []
    preview = []
    for i, row in enumerate(reader):
        if i >= 10:
            break
        preview.append({h: row.get(h, "") for h in headers})

    delim_name = next((name for name, ch in DELIMITER_NAMES.items() if ch == delim_char), "comma")

    return {
        "encoding": detected_encoding,
        "delimiter": delim_name,
        "detected_header_row": header_idx,
        "total_rows": len(lines) - header_idx - 1,
        "headers": headers,
        "preview": preview,
        "format": "csv",
    }


def _profile_column(header: str, rows: list[dict[str, Any]], outlier_bounds: tuple[float, float] | None = None) -> ColumnProfile:
    values = [row.get(header, "") for row in rows]
    normalized = [_normalize_missing(value) for value in values]
    present = [value for value in normalized if value != ""]
    detected_type = _detect_type(present)
    unique_values = len(set(map(str, present)))
    total = len(values)
    missing = len(values) - len(present)
    distribution_pct = round((total - missing) / total * 100, 2) if total > 0 else 0

    profile = ColumnProfile(
        name=header,
        detected_type=detected_type,
        total_rows=total,
        missing=missing,
        unique_values=unique_values,
        examples=list(dict.fromkeys(map(str, present)))[:8],
    )
    profile.distribution_pct = distribution_pct

    if detected_type == "number":
        numeric_values = _to_float_column(present)
        invalid_type_count = len(present) - len([v for v in numeric_values if v is not None])
        profile.invalid_type_count = invalid_type_count
        numeric_values = [value for value in numeric_values if value is not None]
        _add_numeric_stats(profile, numeric_values, outlier_bounds=outlier_bounds)
    else:
        _add_format_groups(profile, present)
        _add_value_distribution(profile, present)

    return profile


def _normalize_missing(value: Any) -> str:
    """Return "" when a value is a missing token, otherwise its stripped text."""
    if is_missing(value):
        return ""
    return str(value).strip()


def _detect_type(values: list[str]) -> str:
    if not values:
        return "text"
    dates = sum(_looks_like_date(str(value)) for value in values)
    numbers = sum(1 for value in _to_float_column(values) if value is not None)
    booleans = sum(str(value).strip().lower() in {"si", "sí", "no", "true", "false", "0", "1"} for value in values)

    total = len(values)
    # AP-02: fecha antes que número, para que "20240101" no se pierda como número.
    if dates / total >= TYPE_THRESHOLD:
        return "date"
    if numbers / total >= TYPE_THRESHOLD:
        return "number"
    if booleans / total >= TYPE_THRESHOLD:
        return "boolean"
    return "text"


def _classify_numeric_separators(values: list[str]) -> str | None:
    """Decide by majority how separators are used across a column (AP-01).

    Returns:
      - "miles": comma groups thousands (45,000 / 1,234,567) per spec.
      - "decimal": comma is the decimal separator (3,5 / 12,34).
      - "dot_miles": dots group thousands (1.234.567); only strong evidence
        (>=2 groups) counts so decimals like 0.500 stay safe.
      - None: ambiguous, fall back to the single-value heuristic.

    The mode that appears strictly more often than the others wins.
    """
    comma_miles = 0
    comma_decimal = 0
    dot_miles = 0
    for value in values:
        s = str(value).strip()
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", s):
            comma_miles += 1
        elif re.fullmatch(r"-?\d+,\d{1,2}", s):
            comma_decimal += 1
        elif re.fullmatch(r"-?\d{1,3}(\.\d{3}){2,}", s):
            dot_miles += 1
    scores = {"miles": comma_miles, "decimal": comma_decimal, "dot_miles": dot_miles}
    top = max(scores, key=scores.get)
    if scores[top] == 0:
        return None
    ordered = sorted(scores.values(), reverse=True)
    if ordered[0] > ordered[1]:
        return top
    return None


def _to_float(value: Any, separator_mode: str | None = None) -> float | None:
    """Parse a value into a float.

    `separator_mode` resolves ambiguous separator usage using the column
    context (AP-01):
      - "miles": comma is a thousands separator (45,000 -> 45000)
      - "decimal": comma is the decimal separator (3,5 -> 3.5)
      - "dot_miles": dots are a thousands separator (1.234.567 -> 1234567)
      - None: the last separator wins (locale heuristic); comma-only values
        keep the comma as decimal separator for backward compatibility.
    """
    s = str(value).strip().replace("\u00a0", "")
    if not s:
        return None
    try:
        if separator_mode == "miles":
            s = s.replace(",", "")
        elif separator_mode == "decimal":
            s = s.replace(".", "").replace(",", ".")
        elif separator_mode == "dot_miles":
            s = s.replace(".", "")
            if "," in s:
                s = s.replace(",", ".")
        elif "," in s and "." in s:
            if s.rfind(".") > s.rfind(","):
                s = s.replace(",", "")
            else:
                s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except ValueError:
        return None


def _to_float_column(values: list[str]) -> list[float | None]:
    """Parse all values of a column, resolving separator ambiguity by majority."""
    separator_mode = _classify_numeric_separators(values)
    return [_to_float(value, separator_mode=separator_mode) for value in values]


def _numeric_outlier_bounds(values: list[float]) -> tuple[float, float] | None:
    """Devuelve (low, high) del rango IQR (1.5x) o None si no hay suficientes
    valores numéricos (<4) o dispersión (IQR=0).

    CL-10: límites de referencia compartidos entre el perfilado original y el
    re-perfilado post-limpieza (el after NO recalcula el IQR sobre datos
    imputados, para que las acciones correctivas no degraden la exactitud).
    """
    if len(values) < 4:
        return None
    sorted_values = sorted(values)
    q1 = statistics.median(sorted_values[: len(sorted_values) // 2])
    q3 = statistics.median(sorted_values[(len(sorted_values) + 1) // 2 :])
    iqr = q3 - q1
    if iqr == 0:
        return None
    return (q1 - 1.5 * iqr, q3 + 1.5 * iqr)


def _outlier_row_indices(rows: list[dict[str, Any]], column: str) -> list[int]:
    """Return 0-based row indices flagged as numeric outliers (CL-03).

    Reusa la MISMA definición de outliers que `_add_numeric_stats` (IQR): requiere
    >=4 valores numéricos y dispersión (IQR > 0); los tokens missing se descartan.
    """
    values = [_normalize_missing(row.get(column, "")) for row in rows]
    parsed = _to_float_column(values)
    numeric = [(i, v) for i, v in enumerate(parsed) if v is not None]
    bounds = _numeric_outlier_bounds([v for _, v in numeric])
    if bounds is None:
        return []
    low, high = bounds
    return [i for i, v in numeric if v < low or v > high]


_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d", "%d%m%Y", "%m%d%Y")


def _looks_like_date(value: str) -> bool:
    for fmt in _DATE_FORMATS:
        try:
            datetime.strptime(value.strip(), fmt)
            return True
        except ValueError:
            continue
    return False


def _add_value_distribution(profile: ColumnProfile, values: list[str]) -> None:
    if not values:
        return
    from collections import Counter
    freq = Counter(str(v).strip() for v in values if str(v).strip())
    n = len(values)
    profile.value_distribution = [
        {"value": val, "freq": count, "pct": round(count / n * 100, 2)}
        for val, count in freq.most_common()
    ]


def _add_numeric_stats(profile: ColumnProfile, values: list[float], outlier_bounds: tuple[float, float] | None = None) -> None:
    if not values:
        return
    profile.min_value = round(min(values), 4)
    profile.max_value = round(max(values), 4)
    profile.mean = round(statistics.fmean(values), 4)
    profile.median = round(statistics.median(values), 4)

    # CL-10: si se reciben límites congelados (del dataset original) se usan
    # tal cual; imputar/rellenar no debe recalcular el IQR sobre datos ya
    # corregidos y convertir valores normales en outliers.
    if outlier_bounds is None:
        outlier_bounds = _numeric_outlier_bounds(values)
        if outlier_bounds is None:
            # AP-05: sin dispersión (o <4 valores) no hay outliers calculables.
            profile.outlier_analysis_skipped = True
            return
    low, high = outlier_bounds
    outliers = [value for value in values if value < low or value > high]
    profile.outliers = len(outliers)
    profile.outlier_examples = outliers[:8]


def _add_format_groups(profile: ColumnProfile, values: list[str]) -> None:
    groups: dict[str, dict[str, int]] = {}
    for value in values:
        key = " ".join(str(value).strip().lower().split())
        variant = str(value)
        groups.setdefault(key, {}).setdefault(variant, 0)
        groups[key][variant] += 1

    for variants in groups.values():
        if len(variants) > 1:
            canonical = max(variants, key=variants.get)
            sorted_variants = sorted(variants)
            profile.format_groups.append({"canonical": canonical, "variants": sorted_variants})
            # AP-06: format_issues cuenta FILAS afectadas (las que se desvían de
            # la forma más frecuente), no variantes distintas.
            profile.format_issues += sum(variants.values()) - variants[canonical]


def _count_duplicate_rows(headers: list[str], rows: list[dict[str, Any]], key_columns: list[str] | None = None) -> int:
    """Count full-row duplicates, optionally using only key_columns for comparison.

    When key_columns is None, compares ALL columns (legacy behavior).
    When key_columns has values, builds the key from those columns only.
    Both use the SAME normalized signature (strip + lower + remove accents),
    so detection, removal and the diagnostic agree (DU-01/DU-02).
    """
    seen: set[tuple[str, ...]] = set()
    duplicates = 0
    compare_headers = key_columns if key_columns else headers
    for row in rows:
        key = tuple(normalize_for_comparison(row.get(h, "")) for h in compare_headers)
        if key in seen:
            duplicates += 1
        else:
            seen.add(key)
    return duplicates


def _quality_scores(columns: list[ColumnProfile], duplicate_rows: int, row_count: int, column_count: int) -> dict[str, float]:
    total_cells = max(row_count * column_count, 1)
    missing = sum(column.missing for column in columns)
    format_issues = sum(column.format_issues for column in columns)
    outliers = sum(column.outliers for column in columns)
    type_errors = sum(column.invalid_type_count for column in columns)

    completeness = _score_from_ratio(missing, total_cells)
    consistency = _score_from_ratio(format_issues, total_cells)
    # AP-03: exactitud estructural = 1 - (errores de tipo + outliers) / celdas.
    accuracy = _score_from_ratio(outliers + type_errors, total_cells)
    uniqueness = _score_from_ratio(duplicate_rows, max(row_count, 1))
    # AP-07: overall ponderado (los pesos suman 1.0).
    scores = {
        "completeness": completeness,
        "consistency": consistency,
        "accuracy": accuracy,
        "uniqueness": uniqueness,
    }
    scores["overall"] = round(sum(scores[key] * QUALITY_WEIGHTS[key] for key in QUALITY_WEIGHTS), 2)
    return scores


def _score_from_ratio(problem_count: int, total: int) -> float:
    return round(max(0, 100 - (problem_count / max(total, 1) * 100)), 2)


def _recommendations(columns: list[ColumnProfile], duplicate_rows: int) -> list[dict[str, str]]:
    recommendations: list[dict[str, str]] = []
    for column in columns:
        if column.missing:
            recommendations.append(
                {
                    "priority": "Alta",
                    "message": f"Validar {column.missing} valores faltantes en '{column.name}' antes de tomar decisiones.",
                }
            )
        if column.format_issues:
            recommendations.append(
                {
                    "priority": "Media",
                    "message": f"Estandarizar variantes de escritura en '{column.name}' para evitar categorías fragmentadas.",
                }
            )
        if column.outliers:
            recommendations.append(
                {
                    "priority": "Media",
                    "message": f"Revisar {column.outliers} valores atipicos en '{column.name}' con la fuente original.",
                }
            )
    if duplicate_rows:
        recommendations.append(
            {
                "priority": "Alta",
                "message": f"Evaluar {duplicate_rows} filas duplicadas; eliminarlas solo si no representan eventos reales.",
            }
        )
    if not recommendations:
        recommendations.append({"priority": "Baja", "message": "No se detectaron problemas críticos en el diagnóstico automatico."})
    return recommendations


def _log_entry(column: str, action: str, reason: str, result: str) -> dict[str, str]:
    return {
        "timestamp": datetime.now(timezone(timedelta(hours=-5))).isoformat(timespec="seconds"),
        "column": column,
        "action": action,
        "reason": reason,
        "result": result,
    }


def _imputation_value(rows: list[dict[str, Any]], column: str, method: str, custom_value: Any = None) -> Any:
    present = [_normalize_missing(row.get(column, "")) for row in rows]
    present = [value for value in present if value != ""]
    if method == "custom":
        return "" if custom_value is None else custom_value
    if not present:
        return ""
    numeric = _to_float_column(present)
    numeric = [value for value in numeric if value is not None]
    if method == "mean" and numeric:
        return round(statistics.fmean(numeric), 4)
    if method == "median" and numeric:
        return round(statistics.median(numeric), 4)
    counts = Counter(map(str, present))
    return counts.most_common(1)[0][0]


def _apply_fill_cells(rows: list[dict[str, Any]], headers: list[str], column: str,
                      target_set: set[int] | None, is_missing_value: Callable[[Any], bool],
                      new_value: Any, action_label: str, ai_reason: str) -> tuple[int, list[dict[str, Any]]]:
    """Rellena celdas de 'column' que cumplen `is_missing_value`, respetando target_rows.

    CL-06: implementación ÚNICA compartida por impute_missing / fill_missing / fill_empty.
    Devuelve (celdas cambiadas, entradas de changelog).
    """
    changed = 0
    entries: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        if target_set is not None and idx not in target_set:
            continue
        old_val = row.get(column, "")
        if not is_missing_value(old_val):
            continue
        row_id = row.get("id", row.get("ID", row.get(headers[0], "?")))
        entries.append({
            "action": action_label,
            "column": column,
            "reason": ai_reason,
            "changes": [{"row": str(row_id), "column": column, "old": str(old_val) or "(vacío)", "new": str(new_value)}],
        })
        row[column] = new_value
        changed += 1
    return changed, entries


def _standardize_text(value: str, mode: str) -> str:
    normalized = " ".join(str(value).strip().split())
    if mode == "upper":
        return normalized.upper()
    if mode == "lower":
        return normalized.lower()
    return normalized.title()


def _clean_filename(filename: str) -> str:
    """Return the analysis filename for the cleaned CSV.

    The cleaned content is always serialized as CSV, so the extension must
    always be .csv regardless of the original file format. Reusing the
    original extension (e.g. .xlsx) would make load_dataset try to read the
    CSV bytes as an XLSX archive and fail.
    """
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    return f"{stem}_limpio.csv"


def _conclusión(analysis: dict[str, Any]) -> str:
    overall = analysis["scores"]["overall"]
    if overall >= 95:
        return f"El dataset alcanza una calidad general de {overall}% y esta listo para análisis descriptivo, manteniendo la documentación de decisiones aplicada."
    if overall >= 80:
        return f"El dataset alcanza una calidad general de {overall}%. Es utilizable, pero conviene revisar los riesgos pendientes antes de decisiones finales."
    return f"El dataset alcanza una calidad general de {overall}%. Se recomienda continuar la depuración antes de usarlo para toma de decisiones."


def _executive_summary(analysis: dict[str, Any]) -> str:
    scores = analysis["scores"]
    return (
        f"Se analizo el dataset '{analysis['filename']}' con {analysis['row_count']} filas y "
        f"{analysis['column_count']} columnas. La calidad general calculada fue de "
        f"{scores['overall']}%, considerando completitud, consistencia, exactitud estructural "
        "y unicidad. Los hallazgos deben interpretarse como diagnóstico técnico inicial y "
        "validarse con reglas de negocio antes de ejecutar cambios definitivos."
    )


def _cleaning_resumen(before: dict[str, Any], after: dict[str, Any], actions: list[dict[str, Any]]) -> str:
    total_actions = len(actions)
    improvements = []
    dims = [("completitud", "completeness"), ("consistencia", "consistency"), ("exactitud estructural", "accuracy"), ("unicidad", "uniqueness")]
    for label, key in dims:
        b = before["scores"][key]
        a = after["scores"][key]
        if b != a:
            diff = round(a - b, 2)
            sign = "+" if diff > 0 else ""
            improvements.append(f"{label} ({b}% -> {a}%, {sign}{diff}%)")
    improvement_text = ", ".join(improvements) if improvements else "sin cambios significativos"
    return (
        f"Se ejecutó un proceso secuencial de limpieza sobre el dataset '{before['filename']}', "
        f"compuesto por {before['row_count']} registros y {before['column_count']} columnas. "
        f"Se documentaron {total_actions} acciones de limpieza. "
        f"La calidad general paso de {before['scores']['overall']}% a {after['scores']['overall']}%. "
        f"Las dimensiones con mejoras: {improvement_text}. "
        "Cada decisión quedo registrada para facilitar auditoria, mantenimiento y reutilización."
    )


def _risk_list(analysis: dict[str, Any]) -> list[str]:
    risks = []
    if analysis["scores"]["completeness"] < 95:
        risks.append("Persisten valores faltantes que pueden sesgar indicadores.")
    if analysis["scores"]["consistency"] < 95:
        risks.append("Persisten inconsistencias de formato que pueden fragmentar categorías.")
    if analysis["scores"]["accuracy"] < 95:
        risks.append("Persisten outliers que requieren validación con la fuente original.")
    if analysis["duplicate_rows"]:
        risks.append("Persisten duplicados que deben evaluarse segun la unidad de análisis.")
    if not risks:
        risks.append("No se identifican riesgos críticos en el diagnóstico posterior a la limpieza.")
    return risks


def generate_audit_log(changelog: list[dict[str, Any]], filename: str = "dataset") -> str:
    """Generate a Markdown changelog documenting every cell-level change made during cleaning.

    Returns a structured document suitable for auditing, regulatory compliance,
    or reproducibility purposes.
    """
    lines: list[str] = []
    lines.append(f"# Bitácora de Cambios - {filename}")
    lines.append("")
    lines.append("Documento de auditoría que registra cada modificación realizada sobre el dataset durante el proceso de limpieza.")
    lines.append("")
    lines.append(f"**Total de acciones registradas:** {len(changelog)}")
    lines.append("")

    if not changelog:
        lines.append("_No se realizaron cambios en el dataset._")
        lines.append("")
        return "\n".join(lines)

    lines.append("---")
    lines.append("")

    for idx, entry in enumerate(changelog, 1):
        action = entry.get("action", "Acción desconocida")
        column = entry.get("column", "?")
        reason = entry.get("reason", "Sin justificación registrada.")
        changes = entry.get("changes", [])

        lines.append(f"## {idx}. {action}")
        lines.append("")
        lines.append(f"- **Columna afectada:** {column}")
        lines.append(f"- **Justificación:** {reason}")
        lines.append(f"- **Filas modificadas:** {len(changes)}")
        lines.append("")

        if changes:
            lines.append("| Fila | Columna | Valor anterior | Valor nuevo |")
            lines.append("|------|---------|---------------|-------------|")
            for ch in changes:
                row_id = str(ch.get("row", "?"))
                col = str(ch.get("column", "?"))
                old_val = str(ch.get("old", ""))
                new_val = str(ch.get("new", ""))
                lines.append(f"| {row_id} | {col} | {old_val} | {new_val} |")
            lines.append("")

        lines.append("---")
        lines.append("")

    lines.append("## Resumen")
    lines.append("")
    lines.append("Documento generado automáticamente por AuditData AI.")
    lines.append(f"- **Archivo original:** {filename}")
    lines.append(f"- **Total de acciones:** {len(changelog)}")
    total_changes = sum(len(e.get("changes", [])) for e in changelog)
    lines.append(f"- **Total de celdas modificadas:** {total_changes}")
    lines.append("")
    lines.append("_Este documento debe conservarse como evidencia del proceso de limpieza de datos._")
    lines.append("")

    return "\n".join(lines)
